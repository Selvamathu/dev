from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import pandas as pd
import json
import datetime
import sys
from dataclasses import dataclass
from builtins import str
from _testmultiphase import Str

# To read excel to DataFrame object. 
# First row is considered as Header
df = pd.read_excel("Mara.XLSX")

# It shows all excel data
print(df)

# Shows all header
# print(df.columns.ravel())

# To print only a particular column data
# print(df["MATNR"].tolist())


# to remove a column of excel data frame object
# df.drop(columns = "MANDT", inplace = True)
# print(df)

# To create new excel
# workbook = Workbook()
# sheet = workbook.active
# 
# sheet["A1"] = "Hello"
# sheet["A2"] = "World"
# 
# workbook.save(filename= "hello_world.xlsx")

# To open an excel sheet
# workbook = load_workbook("Mara.XLSX")
# To get all sheet names
# print(workbook.sheetnames)

# To get active sheet
# sheet = workbook.active
# print(sheet.title)

# to read first cell
# print(sheet["A1"].value)
# print(sheet.cell(row = 1, column = 1).value)

# print(sheet["A1:C3"])
# print(sheet["A"])
# print(sheet["A:B"])

# Get all cells from row 5
# print(sheet[5])

# Get all cells from 5 to 7
# print(sheet[5:7])

# Iterate rows
# for row in sheet.iter_rows(min_row = 1, max_row = 2, min_col = 1, max_col =3):
#     print(row)

# otherway

# for row in sheet.rows:
#     print(row)
    
# Iterate columns
# for column in sheet.iter_cols( min_row = 1, max_row = 2, min_col = 1, max_col = 3):
#     print(column)
    
# iterate value from excel
# for value in sheet.iter_rows(min_row = 1, max_row = 2, min_col = 1, max_col =3, values_only = True):
#     print(value)

# Convert excel to dictionary
# products = {}
# 
# for row in sheet.iter_rows(min_row = 2, min_col = 2, max_col = 5, values_only = True):
#     product = { "create_date":row[1].strftime("%d/%m/%Y"),
#                 "create_person":row[3]}
#     product_id = row[0]
#     products[product_id] = product
#     
# print(json.dumps(products))

# to get python version
# print(sys.version)
# print(sys.version_info)


    
    